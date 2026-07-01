import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import datetime

# Create a workbook and select the active sheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt Chart 12 Minggu"

# Ensure grid lines are visible
ws.views.sheetView[0].showGridLines = True

# Palettes
primary_color = "2B6CB0" # Slate Blue
header_fill_color = "EDF2F7" # Light Gray
zebra_fill_color = "F7FAFC" # Off-white
accent_color = "E2E8F0"
aldi_color = "EBF8FF" # Light blue fill
julia_color = "F0FFF4" # Light green fill
sopyan_color = "FFF5F5" # Light red fill
both_color = "FAF5FF" # Light purple fill

# Fonts
font_title = Font(name="Segoe UI", size=16, bold=True, color="1A202C")
font_subtitle = Font(name="Segoe UI", size=10, italic=True, color="4A5568")
font_header = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
font_timeline_header = Font(name="Segoe UI", size=9, bold=True, color="4A5568")
font_data = Font(name="Segoe UI", size=10, color="2D3748")
font_data_bold = Font(name="Segoe UI", size=10, bold=True, color="2D3748")

# Fills
fill_primary = PatternFill(start_color=primary_color, end_color=primary_color, fill_type="solid")
fill_header = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid") # Dark Gray
fill_zebra = PatternFill(start_color=zebra_fill_color, end_color=zebra_fill_color, fill_type="solid")
fill_timeline_header = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")

# Fills for bar chart cells in timeline
fill_bar_aldi = PatternFill(start_color="3182CE", end_color="3182CE", fill_type="solid") # Strong Blue
fill_bar_julia = PatternFill(start_color="38A169", end_color="38A169", fill_type="solid") # Strong Green
fill_bar_sopyan = PatternFill(start_color="E53E3E", end_color="E53E3E", fill_type="solid") # Red
fill_bar_both = PatternFill(start_color="805AD5", end_color="805AD5", fill_type="solid") # Purple

# Borders
thin_side = Side(border_style="thin", color="CBD5E0")
border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=Side(border_style="medium", color="1A202C"))

# Alignments
align_left = Alignment(horizontal="left", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

# Title Block
ws.merge_cells("A1:T1")
ws["A1"] = "GANTT CHART TIMELINE PROYEK KDKMP"
ws["A1"].font = font_title
ws["A1"].alignment = align_left

ws.merge_cells("A2:T2")
ws["A2"] = "Sistem Informasi Komoditas Desa & Keuangan Modern - Estimasi Jadwal Kerja 12 Minggu (Aldi, Julia, & Sopyan)"
ws["A2"].font = font_subtitle
ws["A2"].alignment = align_left

ws.row_dimensions[1].height = 25
ws.row_dimensions[2].height = 18
ws.row_dimensions[3].height = 10 # spacing

# Table Headers
headers = [
    "Task ID", "Nama Aktivitas", "Fase / Kategori", "Penanggung Jawab",
    "Durasi (Minggu)", "Progress"
]

ws.row_dimensions[4].height = 28

# Write task info headers
for col_idx, header in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_header

# Timeline Headers (Minggu 1 sampai 12)
ws.merge_cells(start_row=4, start_column=7, end_row=4, end_column=18)
ws.cell(row=4, column=7, value="TIMELINE PENGERJAAN (MINGGU 1 - MINGGU 12)").alignment = align_center
ws.cell(row=4, column=7).font = font_header
ws.cell(row=4, column=7).fill = fill_primary
ws.cell(row=4, column=7).border = border_header

# Week Labels (Row 5)
ws.row_dimensions[5].height = 20
for week in range(1, 13):
    col_idx = 6 + week
    cell = ws.cell(row=5, column=col_idx, value=f"M{week}")
    cell.font = font_timeline_header
    cell.fill = fill_timeline_header
    cell.alignment = align_center
    cell.border = border_all

# Merge task info column headers vertically to align with row 5
for col_idx in range(1, 7):
    ws.merge_cells(start_row=4, start_column=col_idx, end_row=5, end_column=col_idx)
    # Re-apply border to the merged area borders
    for r in [4, 5]:
        ws.cell(row=r, column=col_idx).border = border_all

# Data Rows: (Task ID, Name, Phase, Owner, Start Week, End Week, Duration, Progress)
tasks = [
    ("T1", "Analisis Kebutuhan & Dokumen SRS", "Persiapan", "Julia & Sopyan", 1, 2, 2, 1.0),
    ("T2", "Perancangan Database & Pembuatan ERD", "Persiapan", "Julia", 2, 3, 2, 1.0),
    ("T3", "Inisialisasi Project & Setup Git/Lingkungan", "Setup", "Aldi", 3, 4, 2, 1.0),
    ("T4", "Migrasi DB & Seeding Data Awal", "Backend", "Aldi", 4, 5, 2, 1.0),
    ("T5", "Pengembangan Modul Keuangan & Termin", "Backend", "Julia", 5, 7, 3, 1.0),
    ("T6", "Pengembangan Modul Barter & Komoditas", "Backend", "Julia", 6, 8, 3, 1.0),
    ("T7", "Desain Layout Dashboard & Sidebar UI", "Frontend", "Sopyan", 7, 9, 3, 1.0),
    ("T8", "Integrasi Form Input & Validasi Frontend", "Frontend", "Aldi", 8, 10, 3, 1.0),
    ("T9", "Implementasi Role Middleware & Security", "Backend", "Julia", 9, 10, 2, 1.0),
    ("T10", "Pembuatan Automated Testing Suite (QA)", "Testing", "Sopyan", 10, 11, 2, 1.0),
    ("T11", "Bug Fixing & UI Polish", "Handover", "Aldi & Sopyan", 11, 12, 2, 1.0),
    ("T12", "Finalisasi Laporan & Persiapan Deployment", "Handover", "Semua Tim", 12, 12, 1, 1.0),
]

current_row = 6
for task in tasks:
    ws.row_dimensions[current_row].height = 22
    tid, tname, phase, owner, start, end, duration, progress = task
    
    # Write details
    ws.cell(row=current_row, column=1, value=tid).alignment = align_center
    ws.cell(row=current_row, column=2, value=tname).alignment = align_left
    ws.cell(row=current_row, column=3, value=phase).alignment = align_center
    ws.cell(row=current_row, column=4, value=owner).alignment = align_center
    ws.cell(row=current_row, column=5, value=f"{duration} Minggu").alignment = align_center
    
    progress_cell = ws.cell(row=current_row, column=6, value=progress)
    progress_cell.number_format = '0%'
    progress_cell.alignment = align_center
    
    # Styling text cells
    for col_idx in range(1, 7):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.font = font_data
        cell.border = border_all
        if current_row % 2 == 1:
            cell.fill = fill_zebra
            
    # Draw Gantt Bar (Week 1 to 12)
    for week in range(1, 13):
        col_idx = 6 + week
        cell = ws.cell(row=current_row, column=col_idx)
        cell.border = border_all
        
        # If the week falls within the task range
        if start <= week <= end:
            if owner == "Aldi":
                cell.fill = fill_bar_aldi
            elif owner == "Julia":
                cell.fill = fill_bar_julia
            elif owner == "Sopyan":
                cell.fill = fill_bar_sopyan
            else:
                cell.fill = fill_bar_both
        else:
            if current_row % 2 == 1:
                cell.fill = fill_zebra

    current_row += 1

# Add Legend
current_row += 1
ws.cell(row=current_row, column=2, value="Legenda Penanggung Jawab:").font = font_data_bold

current_row += 1
ws.cell(row=current_row, column=2, value="Aldi Burung (PM / Backend / Setup)").font = font_data
ws.cell(row=current_row, column=3).fill = fill_bar_aldi
ws.cell(row=current_row, column=3).border = border_all

current_row += 1
ws.cell(row=current_row, column=2, value="Julia (Analyst / Backend / Security)").font = font_data
ws.cell(row=current_row, column=3).fill = fill_bar_julia
ws.cell(row=current_row, column=3).border = border_all

current_row += 1
ws.cell(row=current_row, column=2, value="Sopyan (Frontend / UI / UX / QA)").font = font_data
ws.cell(row=current_row, column=3).fill = fill_bar_sopyan
ws.cell(row=current_row, column=3).border = border_all

current_row += 1
ws.cell(row=current_row, column=2, value="Bersama (Gabungan Tim)").font = font_data
ws.cell(row=current_row, column=3).fill = fill_bar_both
ws.cell(row=current_row, column=3).border = border_all

# Auto-adjust column widths for info columns
for col in range(1, 7):
    max_len = 0
    col_letter = get_column_letter(col)
    for row in range(4, len(tasks) + 6):
        cell_val = ws.cell(row=row, column=col).value
        if cell_val:
            max_len = max(max_len, len(str(cell_val)))
    ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

# Specific manual adjustment for Task Name
ws.column_dimensions["B"].width = 45

# Set width for week columns to look like a timeline grid
for col in range(7, 19):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].width = 6

# Save the workbook
file_path = "c:/laragon/www/kdkmp/project_management/Gantt_Chart_12_Minggu.xlsx"
wb.save(file_path)
print(f"Excel file created at {file_path}")
